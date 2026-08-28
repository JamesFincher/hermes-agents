#!/usr/bin/env python3
"""Build a formatted workbook from a report spec. Deterministic; no model.

  python3 build_xlsx.py spec.json out.xlsx
spec.json: {"title":..,"sheets":[{"name":..,"columns":[..],"rows":[[..]],
            "money_cols":[2,3],"total_row":true,"notes":".."}]}
Requires openpyxl (pip install openpyxl).
"""
import json, sys

def main():
    if len(sys.argv) < 3:
        print("usage: build_xlsx.py spec.json out.xlsx"); sys.exit(2)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed",
                          "fix": "pip install openpyxl --break-system-packages"})); sys.exit(1)
    spec = json.load(open(sys.argv[1]))
    wb = Workbook(); wb.remove(wb.active)
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    thin = Side(style="thin", color="D1D5DB")
    for sh in spec["sheets"]:
        ws = wb.create_sheet(sh["name"][:31])
        ws.append(sh["columns"])
        for c in range(1, len(sh["columns"]) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        for r in sh["rows"]:
            ws.append(r)
        money_cols = set(sh.get("money_cols", []))
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                if cell.column in money_cols:
                    cell.number_format = '#,##0.00;(#,##0.00)'
                    cell.alignment = Alignment(horizontal="right")
        if sh.get("total_row") and sh["rows"]:
            n = ws.max_row
            ws.append(["Total"] + [""] * (len(sh["columns"]) - 1))
            for c in money_cols:
                L = get_column_letter(c)
                cell = ws.cell(row=n + 1, column=c, value=f"=SUM({L}2:{L}{n})")
                cell.font = Font(bold=True)
                cell.number_format = '#,##0.00;(#,##0.00)'
            ws.cell(row=n + 1, column=1).font = Font(bold=True)
        for c, col in enumerate(sh["columns"], 1):
            width = max(len(str(col)) + 2, *(len(str(r[c-1])) + 2 for r in sh["rows"][:200])) if sh["rows"] else len(str(col)) + 2
            ws.column_dimensions[get_column_letter(c)].width = min(width, 48)
        ws.freeze_panes = "A2"
        if sh.get("notes"):
            ws.append([]); ws.append([sh["notes"]])
    wb.save(sys.argv[2])
    print(json.dumps({"ok": True, "path": sys.argv[2], "sheets": [s["name"] for s in spec["sheets"]]}))


if __name__ == "__main__":
    main()
